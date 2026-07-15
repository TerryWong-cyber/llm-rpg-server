# export_configs.py
import json
import os
import openpyxl

EXCEL_FILE = "game_configs.xlsx"
OUTPUT_JSON = "game_configs.json"


def export():
    if not os.path.exists(EXCEL_FILE):
        print(f"⚠️ 未找到源 Excel 配置文件 '{EXCEL_FILE}'，请确认已创建该文件。")
        return

    print("🚀 开始读取 Excel 配置表并转换为 JSON...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    configs = {}

    # 1. 导出简单结构表 (Characters, Armors, Items)
    for sheet_name in ["Characters", "Armors", "Items"]:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # 获取表头
            headers = [cell.value for cell in sheet[1]]
            sheet_data = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue  # 跳过空行
                row_dict = dict(zip(headers, row))
                # 转换主键为 str 保证查找稳定
                sheet_data[str(row_dict["id"])] = row_dict
            configs[sheet_name.upper()] = sheet_data

    # 2. 导出 Weapons (需特殊反序列化其内部嵌套的 skills JSON 字符串)
    if "Weapons" in wb.sheetnames:
        sheet = wb["Weapons"]
        headers = [cell.value for cell in sheet[1]]
        weapons_data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            row_dict = dict(zip(headers, row))
            # 反序列化技能
            if "skills" in row_dict and row_dict["skills"]:
                try:
                    row_dict["skills"] = json.loads(row_dict["skills"])
                except Exception as e:
                    print(f"❌ 解析武器 {row_dict['name']} 技能 JSON 出错: {e}")
                    row_dict["skills"] = []
            weapons_data[str(row_dict["id"])] = row_dict
        configs["WEAPONS"] = weapons_data

    # 3. 导出 Environments (提取为纯 string 列表)
    if "Environments" in wb.sheetnames:
        sheet = wb["Environments"]
        configs["ENVIRONMENTS"] = [row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[1]]

    # 4. 导出 GlobalConfigs (解析为纯 KV 字典形式，自动根据 type 强制类型转换)
    if "GlobalConfigs" in wb.sheetnames:
        sheet = wb["GlobalConfigs"]
        global_data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            key, val, val_type, _ = row[0], row[1], row[2], row[3]
            # 根据配置的类型做动态转换
            if val_type == "int":
                global_data[key] = int(val)
            elif val_type == "float":
                global_data[key] = float(val)
            else:
                global_data[key] = str(val)
        configs["GLOBAL_CONFIGS"] = global_data

    # 输出合并后的 JSON 配置文件
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(configs, f, ensure_ascii=False, indent=4)

    print(f"✨ 转换完成！静态配置已成功生成到 '{OUTPUT_JSON}'。")


if __name__ == "__main__":
    export()