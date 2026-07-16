from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "game_configs.xlsx"
TARGET = ROOT / "configs" / "catalog" / "game.json"


def export() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    with TARGET.open(encoding="utf-8") as stream:
        existing = json.load(stream)
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    output = {
        "schema_version": existing["schema_version"],
        "characters": _simple_sheet(workbook, "Characters"),
        "armors": _simple_sheet(workbook, "Armors"),
        "items": _simple_sheet(workbook, "Items"),
        "weapons": _merge_weapon_extensions(_weapons(workbook), existing["weapons"]),
        "environments": _environments(workbook),
        "rules": {**existing["rules"], **_rules(workbook)},
        "resources": existing["resources"],
    }
    if "0" in existing["armors"]:
        output["armors"].setdefault("0", existing["armors"]["0"])
    with TARGET.open("w", encoding="utf-8") as stream:
        json.dump(output, stream, ensure_ascii=False, indent=2)
        stream.write("\n")


def _simple_sheet(workbook, sheet_name: str) -> dict:
    if sheet_name not in workbook.sheetnames:
        return {}
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    result = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        item = dict(zip(headers, row, strict=False))
        result[str(item["id"])] = item
    return result


def _weapons(workbook) -> dict:
    result = _simple_sheet(workbook, "Weapons")
    for item in result.values():
        if isinstance(item.get("skills"), str):
            item["skills"] = json.loads(item["skills"])
    return result


def _merge_weapon_extensions(exported: dict, existing: dict) -> dict:
    for weapon_id, weapon in exported.items():
        existing_skills = {
            skill["id"]: skill
            for skill in existing.get(weapon_id, {}).get("skills", [])
        }
        for skill in weapon.get("skills", []):
            extension = existing_skills.get(skill["id"], {})
            for key in ("status_effect", "self_effect", "self_effect_ratio"):
                if key in extension:
                    skill[key] = extension[key]
    return exported


def _environments(workbook) -> list[str]:
    if "Environments" not in workbook.sheetnames:
        return []
    return [row[1] for row in workbook["Environments"].iter_rows(min_row=2, values_only=True) if row[1]]


def _rules(workbook) -> dict:
    if "GlobalConfigs" not in workbook.sheetnames:
        return {}
    result = {}
    for key, value, value_type, *_ in workbook["GlobalConfigs"].iter_rows(min_row=2, values_only=True):
        if not key:
            continue
        result[key] = int(value) if value_type == "int" else float(value) if value_type == "float" else str(value)
    return result


if __name__ == "__main__":
    export()
